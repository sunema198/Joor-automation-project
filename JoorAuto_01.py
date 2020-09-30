import os
import time
import re
import errno
import sys
from pathlib import Path
from itertools import islice
from fuzzywuzzy import fuzz

from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import (
    TimeoutException, NoSuchElementException, NoSuchWindowException
)
from file_namelist import get_linesheets

CWD = Path.cwd()  # Current working dierctory


def get_login_and_password():
    auth_file = CWD / "config"
    username, password = None, None
    if auth_file.exists():
        with auth_file.open() as f:
            for line in f.readlines():
                line = line.strip()
                if "username" in line:
                    username = line.split("=")[1]
                if "password" in line:
                    password = line.split("=")[1]
    if not (username or password):
        username = input("Enter email: ")
        password = input("Enter password: ")
    return username, password


# Login Credentials
login_username, login_password = get_login_and_password()

# Designer id and Linesheet_name given
account_id = input("Please Enter Designer Id: ") or 661209
linesheet = input(
    "Enter linesheets separated by commas: ") or "DE SMET"
input_list = linesheet.split(',')
linesheet_names = [x.strip().lower() for x in input_list]

upload_pic = input(
    "Type p for `Upload pic` or any key for `Upload Photos`: "
).strip().lower()
upload_color_photo = input(
    "Type y if uploading color photos or any key if not: "
).strip().lower()
if upload_color_photo == 'y':
    is_code_present = input(
        "Type y if color code is in image name or any key if not: "
    ).strip().lower()
else:
    is_code_present = 'n'
upload_swatch = input(
    "Type y if `Upload Swatch` or any key if not: "
).strip().lower()

USER_FOLDER = CWD / Path(login_username)
USER_LOGGED_IN = USER_FOLDER.exists()
chrome_options = Options()
chrome_options.add_argument(f"user-data-dir={USER_FOLDER}")
driver = webdriver.Chrome(
    executable_path='chromedriver', options=chrome_options)


# driver.maximize_window()

# Path to the downloaded Imagefolder and Swatchfolder under the name of Designer_id
IMAGE_PATH = CWD / f"{account_id}"
SWATCH_IMAGE_PATH = CWD / f"{account_id}/Swatch"

SCREENSHOT_PATH = CWD / f"Screenshot__{account_id}"
try:
    os.mkdir(SCREENSHOT_PATH)
except OSError as e:
    if e.errno == errno.EEXIST:
        print('Directory not created.')
    else:
        raise


URL = "https://www.jooraccess.com"

EMAIL_ID = "login-name"
PASSWORD_ID = "login-password"
LOGIN_FORM_SUBMIT = "form__submit"
ACCOUNT_EDIT_URL = "https://www.jooraccess.com/admin/accounts/edit/{}"
ADMIN_URL = "https://www.jooraccess.com/{}"
STYLE_URL = "https://www.jooraccess.com/styles"
MULTIPLE_EDIT_STYLE_URL = "https://www.jooraccess.com/styles/index/page%3A1"
STYLE_SEARCH_INPUT_ID = "StyleQuery"
STYLE_SEARCH_BUTTON_PATH = "//input[@alt='Submit']"
UPLOAD_SWATCH_BUTTON_PATH = "//a[contains(text(),'Upload Swatch')]"
UPLOAD_PHOTOS_BUTTON_PATH = "//a[contains(text(),'Upload Photos')]"
UPLOAD_FILES_INPUT_ID = "j-style-file-upload-btn"
UPLOAD_FILES_INPUT_PATH = '//input[@name="files[]"]'
POPUP_FILE_SAVE_BUTTON_ID = "j-import-orders-btn"
MAIN_SAVE_BUTTON = "//a[@class='button button-1 float_left j-form-submit']"

pattern = re.compile(r"[^A-Za-z0-9]+")
two_no_pattern = re.compile(r"[^A-Za-z0-9]{2}")

special_chars = re.compile(r'-|_|/|:')
remove_chars = re.compile(r'-|_|/|:| |front|back|BACK|FRONT')

TAB_NOS = 10


class Joor:
    def __init__(self, username, password):
        self.username = username
        self.password = password
        self.wait = WebDriverWait(driver, 25)
        linesheets, large_photos = get_linesheets(IMAGE_PATH)
        self.linesheets = linesheets
        self.large_photos = large_photos
        self.not_uploaded = []
        self.absent_style_sheets = set()
        self.code_not_found = set()
        self.uploaded = []
        self.tab_photos = {}
        self.open_tabs = []
        self.linesheet_name = linesheet_names
        self.upload_pic = True if upload_pic == 'p' else False
        self.upload_swatch = True if upload_swatch == 'y' else False
        self.color_upload = True if upload_color_photo == 'y' else False
        self.is_code_present = True if is_code_present == 'y' else False

    def login(self):
        driver.get(URL)
        driver.find_element_by_id(EMAIL_ID).send_keys(self.username)
        driver.find_element_by_id(PASSWORD_ID).send_keys(self.password)
        driver.find_element_by_class_name(LOGIN_FORM_SUBMIT).click()
        self.wait.until(EC.visibility_of_element_located(
            (By.CLASS, "admin bg_light")))

    def get_designer_link(self, columns, email=None):
        data = []
        for element in columns:
            if email and email in element.get_text():
                cols = element.find_next("td", class_="actions")
                a1 = cols.find_all("a")
                data.append(a1[1])
            elif "os+" in element.get_text():
                cols = element.find_next("td", class_="actions")
                a1 = cols.find_all("a")
                data.append(a1[1])
            elif "jooraccess.com" in element.get_text():
                cols = element.find_next("td", class_="actions")
                a2 = cols.find_all("a")
                data.append(a2[1])
        return data

    def got_to_designer_page(self):
        account_url = ACCOUNT_EDIT_URL.format(str(account_id))
        driver.get(account_url)
        soup = BeautifulSoup(driver.page_source, features="html.parser")
        email_table = soup.find_all("table")[-1]
        columns = email_table.find_all("td")
        data = self.get_designer_link(columns)
        while not data:
            email = input(
                """Admin link not found. Enter new email to login or press q to quit: """
            ).strip()
            if email == "q":
                sys.exit(0)
            data = self.get_designer_link(columns, email)

        if data[0].get("href"):
            admin_url = ADMIN_URL.format(str(data[0].get("href")))
            driver.get(admin_url)
            driver.get(STYLE_URL)
            self.wait.until(
                EC.visibility_of_element_located(
                    (By.ID, STYLE_SEARCH_INPUT_ID))
            )

    def _open_new_tab(self, url, code):
        tab_name = f"{code}__{url}"
        driver.execute_script(
            "window.open('" + url + "', 'tab" + tab_name + "');")
        driver.switch_to.window("tab" + tab_name)
        self.open_tabs.append(tab_name)

    def _search_style(self, code, main_tab):
        tab = driver.current_window_handle
        if not tab == main_tab:
            driver.switch_to.window(main_tab)

        self.wait.until(EC.visibility_of_element_located(
            (By.ID, STYLE_SEARCH_INPUT_ID)))
        search_elem = driver.find_element_by_id(STYLE_SEARCH_INPUT_ID)
        search_elem.clear()
        search_elem.send_keys(code)
        time.sleep(2)
        driver.find_element_by_xpath(STYLE_SEARCH_BUTTON_PATH).click()
        self.wait.until(EC.visibility_of_element_located(
            (By.ID, STYLE_SEARCH_INPUT_ID)))

    def find_linesheets(self, body, code):
        data = set()
        for tr in body.find_all("tr", class_='small'):
            linesheets_ = []
            for td in tr.find_all('td', {'class': 'text-2 no-border notranslate'}):
                a = td.find_all('a')
                for el in a:
                    ele = el.get_text().lower()
                    linesheets_.append(ele)
            header = tr.find_all('strong', class_='text-1')
            if header:
                text = header[0].get_text()
                if re.search(r'{}[\)| |-|:|\s]+'.format(code), text, re.IGNORECASE):
                    if any(x in linesheets_ for x in self.linesheet_name):
                        a_list = tr.find_all("a", string="Edit")
                        for al in a_list:
                            data.add(al)
        return list(data)

    def _go_to_edit_style(self, code):
        style_soup = BeautifulSoup(driver.page_source, features="html.parser")
        table = style_soup.find_all('table')[-1]
        body = table.find('tbody')
        if not body.find('tr'):
            print(f"Search result not found for {code}.")
            self.code_not_found.add(code)
            return None
        data = self.find_linesheets(body, code)
        while not data:
            self.absent_style_sheets.update(self.linesheet_name)
            linesheet_name = input(
                f"""{linesheet} not found. Enter new linesheet or Press 'n' to 'Skip to Next' press 'q' to 'Quit': """
            ).strip().lower()
            if linesheet_name == "q":
                self.exit()
            if linesheet_name == "n":
                return None
            self.linesheet_name.append(linesheet_name)
            data = self.find_linesheets(body, code)
        return data

    def add_tab_photos(self, photos):
        handle = driver.current_window_handle
        if not isinstance(photos, list):
            photos = [photos]
        if handle in self.tab_photos:
            self.tab_photos[handle].extend(photos)
        else:
            self.tab_photos[handle] = photos

    def upload_swatch_photo(self, photo, code):
        # Uploads single photo in 'Upload swatch'
        actions = ActionChains(driver)
        if photo:
            try:
                swatch_upload = driver.find_element_by_xpath(
                    UPLOAD_SWATCH_BUTTON_PATH)
                actions.move_to_element(swatch_upload).perform()
                swatch_upload.click()
                file_input = self.wait.until(
                    lambda x: x.find_element_by_xpath(
                        '//input[@name="files[]"]')
                )
                file_input.send_keys(str(photo))
                self.wait.until(
                    EC.visibility_of_element_located(
                        (By.XPATH, "//a[@id='j-import-orders-btn']")
                    )
                ).click()
            except (NoSuchElementException, TimeoutException):
                print("Swatch Image not Uploaded")
                self.not_uploaded.append(photo)
        else:
            print("Swatch image of {} not found.".format(code))

    def upload_photo(self, photo):
        # Uploads single photo in 'Upload pic'
        actions = ActionChains(driver)
        if photo:
            try:
                upload_btn = driver.find_element_by_xpath(
                    "//a[contains(text(),'Upload Pic')]")
                actions.move_to_element(upload_btn).perform()
                upload_btn.click()
                file_input = self.wait.until(
                    lambda x: x.find_element_by_xpath(
                        '//input[@name="files[]"]')
                )
                file_input.send_keys(str(photo))
                self.wait.until(
                    EC.visibility_of_element_located(
                        (By.XPATH, "//a[@id='j-import-orders-btn']")
                    )
                ).click()
                self.add_tab_photos(photo)
            except (NoSuchElementException, TimeoutException):
                print("Image not Uploaded")
                self.not_uploaded.append(photo)
        else:
            print("Image for Upload pic not found.")

    def _upload_photos(self, photos):
        # uploads multiple photos by clicking 'Upload Photos' button
        scroll_to_middle = (
            "var viewPortHeight = Math.max(document.documentElement.clientHeight, window.innerHeight || 0);"
            + "var elementTop = arguments[0].getBoundingClientRect().top;"
            + "window.scrollBy(0, elementTop-(viewPortHeight/2));"
        )
        if photos:
            try:
                element = driver.find_element_by_xpath(
                    UPLOAD_PHOTOS_BUTTON_PATH)
                driver.execute_script(scroll_to_middle, element)
                element.click()

                photo_input = self.wait.until(
                    lambda x: x.find_element_by_xpath(UPLOAD_FILES_INPUT_PATH)
                )
                photos_ = " \n ".join(map(str, reversed(photos)))
                photo_input.send_keys(photos_)
            except (NoSuchElementException, TimeoutException):
                print("Images not Uploaded")
                self.not_uploaded.extend(photos)
            else:
                self.add_tab_photos(photos)
        else:
            print("Images for Upload Photos not found.")

    def swatch_common_photo(self, code):
        for photo in SWATCH_IMAGE_PATH.iterdir():
            if code in photo.stem:
                return photo

    def tab_group(self, data, size=TAB_NOS):
        it = iter(data)
        for i in range(0, len(data), size):
            yield {key: data[key] for key in islice(it, size)}

    def uploader(self):
        main_tab = driver.current_window_handle
        for data in self.tab_group(self.linesheets):
            for code, photos in data.items():
                tab_name = code
                # self._open_new_tab(code)
                self._search_style(code, main_tab)
                try:
                    if not self.color_upload:
                        edit_url = self._go_to_edit_style(code)
                        if edit_url:
                            edit_link = edit_url[0].get("href")
                            editstyle_url = URL + str(edit_link)
                            self._open_new_tab(editstyle_url, tab_name)
                            # driver.get(editstyle_url)
                            if self.upload_swatch:
                                swatch_photos = self.swatch_common_photo(code)
                                if swatch_photos:
                                    self.upload_swatch_photo(
                                        swatch_photos, code)
                            if self.upload_pic:
                                first = self.get_first(photos)
                                if first:
                                    self.upload_photo(first)
                                    photos.remove(first)
                            self._upload_photos(photos)
                        else:
                            print("Linesheet of {} is not found.".format(code))
                    else:
                        edit_url = self._go_to_edit_style(code)
                        if edit_url:
                            if len(edit_url) > 1:
                                for edit_href in edit_url:
                                    edit_link = edit_href.get('href')
                                    editstyle_url = URL + str(edit_link)
                                    tab_name = f"{code}__{edit_link}"
                                    self._open_new_tab(editstyle_url, tab_name)
                                    # driver.get(editstyle_url)
                                    self._upload_color_photos(code, photos)
                            else:
                                edit_link = edit_url[0].get("href")
                                editstyle_url = URL + str(edit_link)
                                self._open_new_tab(editstyle_url, tab_name)
                                # driver.get(editstyle_url)
                                self._upload_color_photos(code, photos)
                        else:
                            print("Linesheet of {} is not found.".format(code))
                except NoSuchElementException:
                    print("Images failed to Upload")
            self.save_photos()
            driver.switch_to.window(main_tab)
            self.open_tabs = []

    def save_photos(self):
        # Saves photos by switching on each tabes
        for tab_name in self.open_tabs:
            driver.switch_to.window("tab" + tab_name)
            print("TAB: ", tab_name)
            if 'MULTIPLE_EDIT_STYLE_URL' in driver.current_url:
                driver.close()
            else:
                max_attempts = 3
                attempt = 1
                try:
                    popup = driver.find_element_by_xpath(
                        "//div[@id='popupContent']")
                    if popup.is_displayed():
                        try:
                            while attempt < max_attempts:
                                try:
                                    save_btn = self.wait.until(
                                        EC.element_to_be_clickable(
                                            (By.ID, POPUP_FILE_SAVE_BUTTON_ID)
                                        )
                                    )
                                    if save_btn.is_displayed():
                                        save_btn.click()
                                        print("Saved {}".format(tab_name))
                                    else:
                                        pass
                                except TimeoutException:
                                    print(
                                        f"saving failed {attempt} time", tab_name)
                                    attempt += 1
                                else:
                                    break
                        except (NoSuchElementException, NoSuchWindowException):
                            print("Go to Save")
                            pass

                except NoSuchElementException:
                    print("Upload Photos Popup not found.")
                except NoSuchWindowException:
                    print("Uploading failed.")
                handle = driver.current_window_handle
                self.uploaded.extend(
                    self.tab_photos.get(handle, []))
                screenshot = SCREENSHOT_PATH
                el = driver.find_element_by_xpath(
                    "//div[@class='page']")
                time.sleep(1)
                el.screenshot(
                    "{}/{}.png".format(screenshot, tab_name))
                actions = ActionChains(driver)
                element = driver.find_element_by_xpath(
                    "//a[contains(text(),'Save')]")
                actions.move_to_element(element).perform()
                element.click()
                driver.close()

    def get_first(self, paths):
        pattern = re.compile(r'(-|_| )')
        # paths.sort(key=lambda x: x.stem)
        first = paths[0]
        last_digits = ["1", "01", "front", "f", '(1)']
        for path in paths:
            end = pattern.split(path.stem)[-1].lower()
            if end in last_digits:
                return path
        end = pattern.split(first.stem)[-1].lower()
        if end not in ['back', 'b']:
            return first

    def exit(self):
        self.generate_report()
        sys.exit(0)

    def generate_report(self):
        wb = Workbook()
        ws = wb.active
        rows = [
            'Uploaded', 'Not Uploaded', 'Stylesheet Not Found', 'Code Not Found', 'Large Photos'
        ]
        for i, r in enumerate(rows):
            ws.cell(column=i + 1, row=1, value=r)

        all_photos = set()
        for _, photos in self.linesheets.items():
            all_photos.update(photos)

        self.uploaded = set(self.uploaded)

        if self.uploaded:
            for i, v in enumerate(self.uploaded):
                ws.cell(column=1, row=i + 2, value=v.name)

        not_uploaded = all_photos - self.uploaded
        self.not_uploaded.extend(list(not_uploaded))
        self.not_uploaded = set(self.not_uploaded)
        if self.not_uploaded:
            for i, v in enumerate(self.not_uploaded):
                value = getattr(v, 'name') if hasattr(v, 'name') else str(v)
                ws.cell(column=2, row=i + 2, value=value)

        if self.absent_style_sheets:
            for i, v in enumerate(self.absent_style_sheets):
                value = str(v)
                ws.cell(column=3, row=i + 2, value=value)

        if self.code_not_found:
            for i, v in enumerate(self.code_not_found):
                value = str(v)
                ws.cell(column=4, row=i + 2, value=value)

        if self.large_photos:
            for i, v in enumerate(self.large_photos):
                value = getattr(v, 'name') if hasattr(v, 'name') else str(v)
                ws.cell(column=5, row=i + 2, value=value)

        file_name = CWD / f"report_{account_id}.xlsx"
        wb.save(file_name)

    def start(self):
        print("Starting upload")
        if USER_LOGGED_IN:
            # self.login()
            self.got_to_designer_page()
        else:
            self.login()
            self.got_to_designer_page()
        self.uploader()
        self.generate_report()
        print('Finished uploading')

    def get_color_photos(self, code, color, color_code, photos):
        matched_photos = []
        for photo in photos:
            name = photo.stem[len(code):]
            name = remove_chars.sub(" ", name).lower().strip()
            if self.is_code_present:
                color_code = special_chars.sub(' ', color_code).strip()
                if len(color_code) > 1:
                    if color_code in name:
                        matched_photos.append(photo)
            else:
                color = special_chars.sub(' ', color).strip()
                if len(color) > 1:
                    tst = fuzz.token_set_ratio(name, color)
                    ratio = fuzz.ratio(name, color)
                    if color in name:
                        matched_photos.append(photo)
                    elif ratio > 80:
                        matched_photos.append(photo)
                    elif fuzz.token_set_ratio(name, color) > 80:
                        matched_photos.append(photo)
        return matched_photos

    def _upload_color_photos(self, code, photos):
        table = driver.find_element(By.ID, "j-style-color-list")
        # get all of the rows in the table
        rows = table.find_elements(By.TAG_NAME, "tr")
        all_photos = []
        for i, row in enumerate(rows[1:]):
            cols = row.find_elements(By.TAG_NAME, "td")
            color, color_code = None, None
            if self.is_code_present:
                color_code_id = f"StyleColor{i}ColorCodeColorCode"
                color_code_input = cols[2].find_element(By.ID, color_code_id)
                color_code = color_code_input.get_attribute("value").lower()
                color_photos = self.get_color_photos(
                    code, color=color, color_code=color_code, photos=photos)
            else:
                color_input = cols[1].find_element(
                    By.CLASS_NAME, "color_name_input")
                color = color_input.get_attribute("value").lower()
                color_photos = self.get_color_photos(
                    code, color=color, color_code=color_code, photos=photos)
            if color_photos:
                if self.upload_swatch:
                    swatch_first = self.get_swatch_photo(
                        code, color=color, color_code=color_code)
                    if swatch_first:
                        self._upload_swatch(cols, swatch_first)
                if self.upload_pic:
                    first = self.get_first(color_photos)
                    if first:
                        self._upload_pic(cols, first)
                        color_photos.remove(first)
                all_photos.extend(color_photos)
        self._upload_photos(all_photos)

    def _upload_pic(self, cols, pic):
        actions = ActionChains(driver)
        try:
            color_upload = cols[4].find_element(
                By.CLASS_NAME, "j-style-color-upload"
            )
            actions.move_to_element(color_upload).perform()
            color_upload.click()
            file_input = self.wait.until(
                lambda x: x.find_element_by_xpath(
                    '//input[@name="files[]"]')
            )
            file_input.send_keys(str(pic))
            self.wait.until(
                EC.visibility_of_element_located(
                    (By.XPATH, "//a[@id='j-import-orders-btn']")
                )
            ).click()

        except (NoSuchElementException, TimeoutException):
            print("Color Image not Uploaded")
            self.not_uploaded.append(pic)
        else:
            self.add_tab_photos(pic)

    def get_swatch_photo(self, code, color, color_code):
        swatch_path = SWATCH_IMAGE_PATH
        for photo in swatch_path.iterdir():
            if code in photo.stem:
                name = photo.stem[len(code):]
                name = remove_chars.sub(" ", name).lower().strip()
                if self.is_code_present:
                    color_code = special_chars.sub(' ', color_code).strip()
                    if len(color_code) > 1:
                        if color_code in name:
                            return photo
                else:
                    color = special_chars.sub(' ', color).strip()
                    ratio = fuzz.ratio(name, color)
                    tst = fuzz.token_set_ratio(name, color)
                    if color in name:
                        return photo
                    elif ratio > 80:
                        return photo
                    elif fuzz.token_set_ratio(name, color) > 80:
                        return photo

    def _upload_swatch(self, cols, swatch_pic):
        actions = ActionChains(driver)
        try:
            color_upload = cols[3].find_element(
                By.CLASS_NAME, "j-style-color-upload"
            )
            actions.move_to_element(color_upload).perform()
            color_upload.click()
            file_input = self.wait.until(
                lambda x: x.find_element_by_xpath(
                    '//input[@name="files[]"]')
            )
            file_input.send_keys(str(swatch_pic))
            self.wait.until(
                EC.visibility_of_element_located(
                    (By.XPATH, "//a[@id='j-import-orders-btn']")
                )
            ).click()
            self.add_tab_photos(swatch_pic)
        except (NoSuchElementException, TimeoutException):
            print("Swatch Color Image not Uploaded")
            pass


automation = Joor(login_username, login_password)
automation.start()
